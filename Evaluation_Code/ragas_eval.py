"""
===========================================================================
  END-TO-END RAG EVALUATION MODULE
  Master's Thesis: LLM-Powered Climate Policy Summariser & Gap Analyser
  University of Europe for Applied Sciences, Potsdam  |  2026
===========================================================================

SCIENTIFIC PURPOSE
-------------------
This module evaluates the full RAG pipeline answering RQ1 and RQ2:

  RQ1: How accurately does the RAG system answer climate policy queries?
  RQ2: How faithful are generated answers to the retrieved evidence?

THREE EVALUATION FRAMEWORKS
-----------------------------
  FRAMEWORK 1 -- RAGAS (Reference-Free, LLM-as-Judge)
    Faithfulness, Answer Relevancy, Context Precision, Context Recall
    Cite: Es et al. (2024, EACL)

  FRAMEWORK 2 -- ROUGE (n-gram Overlap vs Gold Answers)
    ROUGE-1, ROUGE-2, ROUGE-L (F1, Precision, Recall)
    Cite: Lin (2004, ACL)

  FRAMEWORK 3 -- BERTScore (Semantic Similarity vs Gold)
    BERTScore Precision, Recall, F1
    Cite: Zhang et al. (2020, ICLR)

  FRAMEWORK 4 -- HALLUCINATION DETECTION (Novel per-claim analysis)
    Supported / Partial / Unsupported / Fabricated claim counts
    Grounding score, Hallucination rate, Citation accuracy check

ADDITIONAL ANALYSES
--------------------
  Per-category breakdown (factual / comparative / gap_detection)
  Per-difficulty analysis (easy / medium / hard)
  Failure mode taxonomy (6 modes)
  Citation accuracy verification ([doc, p.N] tag checking)

OUTPUTS
--------
  data/eval/ragas_results_{model}.json
  data/eval/ragas_benchmark_table_{model}.csv
  data/eval/ragas_benchmark_table_{model}.xlsx
  data/eval/ragas_per_query_{model}.csv
  data/eval/ragas_failure_analysis_{model}.json
  data/eval/ragas_thesis_section_{model}.txt

INSTALLATION (run once)
------------------------
  pip install ragas datasets openai rouge-score bert-score scipy

USAGE
------
  # Quick test -- 10 questions (~$0.05)
  python ragas_eval.py --n_questions 10 --category factual

  # Full run -- 80 questions (~$0.80-1.20 total)
  python ragas_eval.py --n_questions 80

  # ROUGE + BERTScore only (free -- no RAGAS API calls)
  python ragas_eval.py --n_questions 80 --skip_ragas

  # Skip BERTScore (no model download)
  python ragas_eval.py --n_questions 80 --skip_bertscore

  # Skip hallucination analysis (saves ~50% cost)
  python ragas_eval.py --n_questions 80 --skip_hallucination
===========================================================================
"""
from __future__ import annotations

import argparse, csv, json, logging, os, re, sys, time, warnings
import math
import os
os.environ["RAGAS_DO_NOT_USE_OPENAI_EMBEDDINGS"] = "1"
from collections import Counter, defaultdict
from dataclasses import dataclass, field, asdict
from pathlib import Path
from typing import Optional

logging.getLogger("sentence_transformers").setLevel(logging.ERROR)
logging.getLogger("huggingface_hub").setLevel(logging.ERROR)
os.environ.setdefault("TOKENIZERS_PARALLELISM", "false")
os.environ.setdefault("TRANSFORMERS_VERBOSITY", "error")
warnings.filterwarnings("ignore", message=".*position_ids.*")
warnings.filterwarnings("ignore", message=".*unauthenticated.*")


# =============================================================================
#  SECTION 1 -- NATIVE ROUGE (no external dependency)
# =============================================================================

def _tok(text: str) -> list:
    return [t for t in re.sub(r"[^\w\s]", " ", text.lower()).split() if t]

def _ngrams(toks: list, n: int) -> Counter:
    return Counter(tuple(toks[i:i+n]) for i in range(len(toks)-n+1))

def rouge_n(hyp: str, ref: str, n: int) -> dict:
    h, r = _ngrams(_tok(hyp), n), _ngrams(_tok(ref), n)
    if not r: return {"f": 0.0, "p": 0.0, "r": 0.0}
    ov = sum((h & r).values())
    p  = ov / sum(h.values()) if h else 0.0
    rc = ov / sum(r.values())
    f  = 2*p*rc/(p+rc) if p+rc else 0.0
    return {"f": round(f,4), "p": round(p,4), "r": round(rc,4)}

def rouge_l(hyp: str, ref: str) -> dict:
    h, r = _tok(hyp), _tok(ref)
    if not h or not r: return {"f": 0.0, "p": 0.0, "r": 0.0}
    m, n = len(h), len(r)
    dp = [[0]*(n+1) for _ in range(m+1)]
    for i in range(1,m+1):
        for j in range(1,n+1):
            dp[i][j] = dp[i-1][j-1]+1 if h[i-1]==r[j-1] else max(dp[i-1][j],dp[i][j-1])
    lcs = dp[m][n]
    p = lcs/m; rc = lcs/n
    f = 2*p*rc/(p+rc) if p+rc else 0.0
    return {"f": round(f,4), "p": round(p,4), "r": round(rc,4)}

def compute_rouge(hyp: str, ref: str) -> dict:
    return {"rouge1": rouge_n(hyp,ref,1), "rouge2": rouge_n(hyp,ref,2), "rougeL": rouge_l(hyp,ref)}


# =============================================================================
#  SECTION 2 -- OPENAI HELPER (urllib only -- no SDK needed)
# =============================================================================

def _openai_call(api_key: str, messages: list, max_tokens: int = 600) -> str:
    import urllib.request, urllib.error
    payload = json.dumps({"model":"gpt-4o","messages":messages,
                          "max_tokens":max_tokens,"temperature":0.0}).encode()
    req = urllib.request.Request(
        "https://api.openai.com/v1/chat/completions", data=payload,
        headers={"Authorization":f"Bearer {api_key}","Content-Type":"application/json"},
        method="POST")
    try:
        with urllib.request.urlopen(req, timeout=60) as resp:
            return json.loads(resp.read())["choices"][0]["message"]["content"].strip()
    except urllib.error.HTTPError as e:
        raise RuntimeError(f"OpenAI {e.code}: {e.read().decode()[:200]}")


# =============================================================================
#  SECTION 3 -- HALLUCINATION ANALYSER
# =============================================================================

_EXTRACT_PROMPT = """Extract every distinct factual claim from this answer as a JSON array.
Each claim = one verifiable statement. Return ONLY valid JSON, no other text.

ANSWER:
{answer}

Example output: ["India targets 50% non-fossil energy by 2030.", "This target is conditional."]"""

_VERIFY_PROMPT = """You are a faithfulness evaluator for a climate policy QA system.

CONTEXT (retrieved chunks):
{context}

CLAIM: {claim}

Reply with ONLY: LABEL | one-sentence reason
Labels: SUPPORTED | PARTIAL | UNSUPPORTED | FABRICATED"""


class HallucinationAnalyser:
    def __init__(self, api_key: str):
        self._key = api_key

    def extract_claims(self, answer: str) -> list:
        try:
            raw = _openai_call(self._key,
                [{"role":"user","content":_EXTRACT_PROMPT.format(answer=answer)}], 400)
            raw = re.sub(r"```(?:json)?", "", raw).strip().strip("`")
            claims = json.loads(raw)
            if isinstance(claims, list):
                return [str(c) for c in claims[:12]]
        except Exception:
            pass
        return [s.strip() for s in re.split(r"[.!?]", answer) if len(s.strip())>20][:8]

    def verify(self, claim: str, chunks: list) -> dict:
        ctx = "\n---\n".join(c[:500] for c in chunks[:5])
        try:
            raw = _openai_call(self._key,
                [{"role":"user","content":_VERIFY_PROMPT.format(context=ctx,claim=claim)}], 100)
            parts = raw.split("|",1)
            label = parts[0].strip().upper()
            reason = parts[1].strip() if len(parts)>1 else ""
            if label not in {"SUPPORTED","PARTIAL","UNSUPPORTED","FABRICATED"}:
                label = "UNSUPPORTED"
        except Exception:
            label, reason = "UNSUPPORTED", "verification failed"
        return {"claim":claim,"label":label,"reason":reason}

    def analyse(self, answer: str, chunks: list) -> dict:
        claims   = self.extract_claims(answer)
        verified = [self.verify(c, chunks) for c in claims]
        counts   = Counter(v["label"] for v in verified)
        n        = max(len(verified), 1)
        gs = (counts["SUPPORTED"]*1.0 + counts["PARTIAL"]*0.5
              - counts["FABRICATED"]*0.5) / n
        return {
            "n_claims": len(verified),
            "supported": counts["SUPPORTED"],
            "partial":   counts["PARTIAL"],
            "unsupported": counts["UNSUPPORTED"],
            "fabricated":  counts["FABRICATED"],
            "grounding_score":   round(max(0.0,min(1.0,gs)),4),
            "hallucination_rate": round((counts["UNSUPPORTED"]+counts["FABRICATED"])/n,4),
            "claims": verified,
        }


# =============================================================================
#  SECTION 4 -- CITATION CHECKER
# =============================================================================

class CitationChecker:
    def check(self, answer: str, chunks: list) -> dict:
        found = re.findall(r"\[([A-Za-z0-9_\-\.]+),\s*p\.?\s*(\d+)\]", answer)
        if not found:
            return {"n_citations":0,"correct":0,"wrong_page":0,"hallucinated":0,"accuracy":None,"details":[]}
        doc_pages = defaultdict(set)
        for c in chunks:
            did = c.get("doc_id","")
            pg  = c.get("page_start",0)
            if did: doc_pages[did].add(int(pg))
        correct = wrong_page = hallucinated = 0
        details = []
        for doc_id, page_str in found:
            pg = int(page_str)
            if doc_id not in doc_pages:
                status = "HALLUCINATED"; hallucinated += 1
            elif any(abs(pg-p)<=2 for p in doc_pages[doc_id]):
                status = "CORRECT"; correct += 1
            else:
                status = "WRONG_PAGE"; wrong_page += 1
            details.append({"citation":f"[{doc_id},p.{pg}]","status":status})
        n = len(found)
        return {"n_citations":n,"correct":correct,"wrong_page":wrong_page,
                "hallucinated":hallucinated,"accuracy":round(correct/n,4) if n else None,
                "details":details}


# =============================================================================
#  SECTION 5 -- FAILURE MODE CLASSIFIER
# =============================================================================

def classify_failure(ragas: dict, rouge: dict, hall: dict,
                     cite: dict, retrieval_hit: bool) -> list:
    failures = []
    faith = ragas.get("faithfulness",1.0)
    rel   = ragas.get("answer_relevancy",1.0)
    cprec = ragas.get("context_precision",1.0)
    crec  = ragas.get("context_recall",1.0)
    r1f   = rouge.get("rouge1",{}).get("f",1.0)
    hall_r= hall.get("hallucination_rate",0.0)
    if not retrieval_hit:       failures.append("RETRIEVAL_MISS")
    if cprec < 0.3:             failures.append("LOW_CONTEXT_PRECISION")
    if crec  < 0.3:             failures.append("LOW_CONTEXT_RECALL")
    if faith < 0.5:             failures.append("LOW_FAITHFULNESS")
    if hall_r > 0.3:            failures.append("HIGH_HALLUCINATION")
    if rel   < 0.5:             failures.append("LOW_RELEVANCY")
    if r1f   < 0.2:             failures.append("LOW_LEXICAL_OVERLAP")
    if cite.get("hallucinated",0)>0: failures.append("HALLUCINATED_CITATIONS")
    return failures if failures else ["NO_FAILURE"]


# =============================================================================
#  SECTION 6 -- DATA CLASSES
# =============================================================================

@dataclass
class QueryResult:
    qid: str; category: str; difficulty: str
    question: str; ground_truth: str; answer: str
    contexts:     list  = field(default_factory=list)
    insufficient: bool  = False
    latency_ms:   float = 0.0
    ragas:        dict  = field(default_factory=dict)
    rouge:        dict  = field(default_factory=dict)
    bertscore:    dict  = field(default_factory=dict)
    hallucination:dict  = field(default_factory=dict)
    citations:    dict  = field(default_factory=dict)
    failure_modes:list  = field(default_factory=list)
    tokens:       dict  = field(default_factory=dict)

@dataclass
class EvalSummary:
    n_total: int; n_evaluated: int; n_insufficient: int
    model: str;   top_k: int
    faithfulness:          float = 0.0
    answer_relevancy:      float = 0.0
    context_precision:     float = 0.0
    context_recall:        float = 0.0
    rouge1_f: float = 0.0; rouge2_f: float = 0.0; rougeL_f: float = 0.0
    bertscore_f: float = 0.0; bertscore_p: float = 0.0; bertscore_r: float = 0.0
    avg_hallucination_rate: float = 0.0
    avg_grounding_score:    float = 0.0
    avg_citation_accuracy:  Optional[float] = None
    by_category:    dict = field(default_factory=dict)
    by_difficulty:  dict = field(default_factory=dict)
    failure_counts: dict = field(default_factory=dict)
    estimated_cost_usd: float = 0.0


# =============================================================================
#  SECTION 7 -- RAGAS RUNNER
# =============================================================================

def run_ragas(results: list, api_key: str) -> list:
    os.environ["OPENAI_API_KEY"] = api_key
    try:
        import ragas
        print(f"  RAGAS version: {getattr(ragas,'__version__','?')}")
    except ImportError:
        print("  [ERROR] ragas not installed. Run: pip install ragas datasets openai")
        return results

    try:
        from datasets import Dataset
    except ImportError:
        print("  [ERROR] datasets not installed. Run: pip install datasets")
        return results

    dataset = Dataset.from_dict({
        "question":     [r.question     for r in results],
        "answer":       [r.answer       for r in results],
        "contexts":     [r.contexts[:5] for r in results],
        "ground_truth": [r.ground_truth for r in results],
    })

    metric_keys = ["faithfulness","answer_relevancy",
                   "context_precision","context_recall"]

    print("  Running RAGAS (GPT-4o judge -- ~3 API calls per question)...")
    # Try modern API first, then legacy
    for attempt in [1, 2]:
        try:
            if attempt == 1:
                from ragas import evaluate
                from ragas.metrics import (faithfulness, answer_relevancy,
                                           context_precision, context_recall)
                result = evaluate(dataset, metrics=[faithfulness,
                    answer_relevancy, context_precision, context_recall])
            else:
                from ragas import evaluate
                from ragas.metrics import (AnswerFaithfulness, AnswerRelevancy,
                                           ContextPrecision, ContextRecall)
                result = evaluate(dataset, metrics=[AnswerFaithfulness(),
                    AnswerRelevancy(), ContextPrecision(), ContextRecall()])

            df = result.to_pandas()
            scores = df.to_dict(orient="records")
            for i, r in enumerate(results):
                if i < len(scores):
                    r.ragas = {k: round(float(scores[i].get(k,0) or 0),4)
                               for k in metric_keys if k in scores[i]}
            print(f"  RAGAS complete: {result}")
            return results

        except Exception as e:
            print(f"  RAGAS attempt {attempt} failed: {e}")
            if attempt == 2:
                print("  Continuing without RAGAS scores.")
    return results


# =============================================================================
#  SECTION 8 -- BERTSCORE RUNNER
# =============================================================================

def run_bertscore(results: list) -> list:
    try:
        from bert_score import score as bs
    except ImportError:
        print("  [WARN] bert_score not installed. Run: pip install bert-score")
        return results
    hyps = [r.answer       for r in results]
    refs = [r.ground_truth for r in results]
    print("  Computing BERTScore (distilbert-base-uncased)...")
    try:
        P, R, F = bs(hyps, refs, model_type="distilbert-base-uncased",
                     lang="en", verbose=False, device=None)
        for i, r in enumerate(results):
            r.bertscore = {"precision":round(float(P[i]),4),
                           "recall":   round(float(R[i]),4),
                           "f1":       round(float(F[i]),4)}
        print("  BERTScore complete.")
    except Exception as e:
        print(f"  BERTScore failed: {e}")
    return results


# =============================================================================
#  SECTION 9 -- MAIN EVALUATOR
# =============================================================================

class RAGEvaluator:
    def __init__(self, base_dir, api_key, model="gpt4o", top_k=5,
                 skip_ragas=False, skip_bertscore=False,
                 skip_hallucination=False, verbose=True):
        self.base_dir          = Path(base_dir)
        self.api_key           = api_key
        self.model             = model
        self.top_k             = top_k
        self.skip_ragas        = skip_ragas
        self.skip_bertscore    = skip_bertscore
        self.skip_hallucination= skip_hallucination
        self.verbose           = verbose
        os.chdir(self.base_dir)
        sys.path.insert(0, str(self.base_dir/"LLM_Integration_Code"))

    def _log(self, msg):
        if self.verbose: print(msg)

    def _filter(self, q: dict):
        cats = q.get("countries",[])
        cat  = q.get("category","")
        if cat == "comparative":   return None
        if q.get("ipcc_relevant") and not cats: return {"source":"IPCC_AR6"}
        if len(cats) == 1:         return {"country":cats[0]}
        return None

    def collect(self, questions: list, pipe) -> list:
        self._log("\n  [Step 1] Collecting RAG responses...")
        results = []
        for i, q in enumerate(questions, 1):
            self._log(f"    [{i:>3}/{len(questions)}] {q['id']} {q['question'][:50]}...")
            t0 = time.time()
            try:
                resp = pipe.query(q["question"], filters=self._filter(q),
                                  top_k=self.top_k, save=True)
                lat = (time.time()-t0)*1000
                r = QueryResult(
                    qid=q["id"], category=q["category"], difficulty=q["difficulty"],
                    question=q["question"], ground_truth=q["ground_truth"],
                    answer=resp.answer, contexts=[c.get("text","") for c in resp.retrieved_chunks],
                    insufficient=resp.insufficient, latency_ms=round(lat,1),
                    tokens=resp.token_stats)
            except Exception as e:
                self._log(f"      ERROR: {e}")
                r = QueryResult(qid=q["id"], category=q["category"],
                    difficulty=q["difficulty"], question=q["question"],
                    ground_truth=q["ground_truth"],
                    answer=f"[ERROR: {e}]", insufficient=True)
            results.append(r)
            time.sleep(0.3)
        return results

    def rouge_pass(self, results: list) -> list:
        self._log("\n  [Step 2] Computing ROUGE...")
        for r in results:
            r.rouge = compute_rouge(r.answer, r.ground_truth)
        return results

    def hallucination_pass(self, results: list) -> list:
        if self.skip_hallucination:
            self._log("\n  [Step 5] Hallucination analysis SKIPPED.")
            return results
        self._log("\n  [Step 5] Hallucination analysis (~3-5 API calls/query)...")
        analyser = HallucinationAnalyser(self.api_key)
        checker  = CitationChecker()
        for i, r in enumerate(results, 1):
            self._log(f"    [{i:>3}/{len(results)}] {r.qid}...")
            if r.insufficient or r.answer.startswith("[ERROR"):
                r.hallucination = {"n_claims":0,"grounding_score":0.0,
                    "hallucination_rate":1.0,"supported":0,"partial":0,
                    "unsupported":0,"fabricated":0,"claims":[]}
                r.citations = {"n_citations":0,"accuracy":None,"details":[]}
                continue
            r.hallucination = analyser.analyse(r.answer, r.contexts)
            r.citations     = checker.check(r.answer,
                [{"doc_id":c.split()[-1],"page_start":0} for c in r.contexts])
        return results

    def failure_pass(self, results: list) -> list:
        self._log("\n  [Step 6] Classifying failure modes...")
        for r in results:
            r.failure_modes = classify_failure(r.ragas, r.rouge, r.hallucination,
                r.citations, bool(r.contexts) and not r.insufficient)
        return results

    def aggregate(self, results: list) -> EvalSummary:
        def avg(vals):
            v = [x for x in vals if x is not None and not (isinstance(x, float) and math.isnan(x))]
            return round(sum(v)/len(v),4) if v else 0.0

        valid = [r for r in results if not r.answer.startswith("[ERROR")]
        n_ins = sum(1 for r in results if r.insufficient)
        tokens = sum(r.tokens.get("total_tokens",0) for r in results)

        s = EvalSummary(
            n_total=len(results), n_evaluated=len(valid), n_insufficient=n_ins,
            model=self.model, top_k=self.top_k,
            faithfulness=     avg([r.ragas.get("faithfulness")      for r in valid if r.ragas]),
            answer_relevancy= avg([r.ragas.get("answer_relevancy")  for r in valid if r.ragas]),
            context_precision=avg([r.ragas.get("context_precision") for r in valid if r.ragas]),
            context_recall=   avg([r.ragas.get("context_recall")    for r in valid if r.ragas]),
            rouge1_f=avg([r.rouge.get("rouge1",{}).get("f") for r in valid]),
            rouge2_f=avg([r.rouge.get("rouge2",{}).get("f") for r in valid]),
            rougeL_f=avg([r.rouge.get("rougeL",{}).get("f") for r in valid]),
            bertscore_f=avg([r.bertscore.get("f1")        for r in valid]),
            bertscore_p=avg([r.bertscore.get("precision") for r in valid]),
            bertscore_r=avg([r.bertscore.get("recall")    for r in valid]),
            avg_hallucination_rate=avg([r.hallucination.get("hallucination_rate") for r in valid if r.hallucination]),
            avg_grounding_score=   avg([r.hallucination.get("grounding_score")    for r in valid if r.hallucination]),
            avg_citation_accuracy= avg([r.citations.get("accuracy") for r in valid
                                        if r.citations and r.citations.get("accuracy") is not None]) or None,
            estimated_cost_usd=round(tokens*0.000005,3),
        )

        for cat in ["factual","comparative","gap_detection"]:
            cm = [r for r in valid if r.category==cat]
            if cm:
                s.by_category[cat] = {
                    "n": len(cm),
                    "faithfulness":    avg([r.ragas.get("faithfulness")      for r in cm if r.ragas]),
                    "answer_relevancy":avg([r.ragas.get("answer_relevancy")  for r in cm if r.ragas]),
                    "context_precision":avg([r.ragas.get("context_precision")for r in cm if r.ragas]),
                    "context_recall":  avg([r.ragas.get("context_recall")    for r in cm if r.ragas]),
                    "rouge1_f":        avg([r.rouge.get("rouge1",{}).get("f") for r in cm]),
                    "rougeL_f":        avg([r.rouge.get("rougeL",{}).get("f") for r in cm]),
                    "bertscore_f":     avg([r.bertscore.get("f1") for r in cm if r.bertscore]),
                    "hallucination_rate":avg([r.hallucination.get("hallucination_rate") for r in cm if r.hallucination]),
                    "avg_latency_ms":  avg([r.latency_ms for r in cm]),
                }
        for diff in ["easy","medium","hard"]:
            dm = [r for r in valid if r.difficulty==diff]
            if dm:
                s.by_difficulty[diff] = {
                    "n": len(dm),
                    "faithfulness": avg([r.ragas.get("faithfulness") for r in dm if r.ragas]),
                    "rouge1_f":     avg([r.rouge.get("rouge1",{}).get("f") for r in dm]),
                    "bertscore_f":  avg([r.bertscore.get("f1") for r in dm if r.bertscore]),
                    "hallucination_rate":avg([r.hallucination.get("hallucination_rate") for r in dm if r.hallucination]),
                }

        all_fm = []
        for r in valid: all_fm.extend(r.failure_modes)
        s.failure_counts = dict(Counter(all_fm))
        return s


# =============================================================================
#  SECTION 10 -- OUTPUT WRITERS
# =============================================================================

def print_summary(s: EvalSummary):
    W = 65
    print(f"\n{'='*W}")
    print(f"  END-TO-END RAG EVALUATION -- {s.model.upper()} -- n={s.n_evaluated}/{s.n_total}")
    print(f"{'-'*W}")
    def row(lbl, val, fmt=".4f"):
        print(f"  {lbl:<32}  {val:{fmt}}" if val is not None else f"  {lbl:<32}  N/A")
    print("  RAGAS")
    row("Faithfulness",      s.faithfulness)
    row("Answer Relevancy",  s.answer_relevancy)
    row("Context Precision", s.context_precision)
    row("Context Recall",    s.context_recall)
    print(f"{'-'*W}\n  ROUGE vs Gold")
    row("ROUGE-1 F1", s.rouge1_f); row("ROUGE-2 F1", s.rouge2_f); row("ROUGE-L F1", s.rougeL_f)
    print(f"{'-'*W}\n  BERTScore vs Gold")
    row("BERTScore F1", s.bertscore_f); row("BERTScore P", s.bertscore_p); row("BERTScore R", s.bertscore_r)
    print(f"{'-'*W}\n  Hallucination")
    row("Avg Hallucination Rate", s.avg_hallucination_rate)
    row("Avg Grounding Score",    s.avg_grounding_score)
    row("Citation Accuracy",      s.avg_citation_accuracy or 0)
    print(f"{'-'*W}\n  Failure Modes")
    for mode, count in sorted(s.failure_counts.items(), key=lambda x:-x[1]):
        pct = count/s.n_evaluated*100 if s.n_evaluated else 0
        print(f"  {mode:<35} {count:>3} ({pct:.0f}%)")
    print(f"{'-'*W}")
    print(f"  Insufficient: {s.n_insufficient}/{s.n_total}   Cost: ~${s.estimated_cost_usd:.3f}")
    print(f"{'-'*W}\n  PER-CATEGORY")
    print(f"  {'Category':<18} {'N':>4} {'Faith':>7} {'Rel':>7} {'R1':>7} {'BSCO':>7} {'Hall%':>7}")
    print(f"  {'-'*55}")
    for cat, d in s.by_category.items():
        print(f"  {cat:<18} {d['n']:>4} "
              f"{d.get('faithfulness',0):>7.4f} {d.get('answer_relevancy',0):>7.4f} "
              f"{d.get('rouge1_f',0):>7.4f} {d.get('bertscore_f',0):>7.4f} "
              f"{d.get('hallucination_rate',0)*100:>6.1f}%")
    print(f"{'-'*W}\n  PER-DIFFICULTY")
    for diff, d in s.by_difficulty.items():
        print(f"  {diff:<10} n={d['n']} faith={d.get('faithfulness',0):.4f} "
              f"rouge1={d.get('rouge1_f',0):.4f} hall%={d.get('hallucination_rate',0)*100:.1f}%")
    print(f"{'-'*W}")


def save_results(summary: EvalSummary, results: list, out_dir: Path, model: str):
    out_dir.mkdir(parents=True, exist_ok=True)

    # JSON
    rj = out_dir/f"ragas_results_{model}.json"
    with open(rj,"w",encoding="utf-8") as f:
        json.dump({"summary":asdict(summary),"per_query":[asdict(r) for r in results]},
                  f,indent=2,default=str)
    print(f"  [SAVED] {rj}")

    # Per-query CSV
    rc = out_dir/f"ragas_per_query_{model}.csv"
    fields=["qid","category","difficulty","faithfulness","answer_relevancy",
            "context_precision","context_recall","rouge1_f","rouge2_f","rougeL_f",
            "bertscore_f","hallucination_rate","grounding_score","citation_accuracy",
            "n_citations","latency_ms","insufficient","failure_modes",
            "question","answer_preview"]
    with open(rc,"w",newline="",encoding="utf-8-sig") as f:
        w = csv.DictWriter(f,fieldnames=fields,extrasaction="ignore")
        w.writeheader()
        for r in results:
            w.writerow({
                "qid":r.qid,"category":r.category,"difficulty":r.difficulty,
                "faithfulness":r.ragas.get("faithfulness",""),
                "answer_relevancy":r.ragas.get("answer_relevancy",""),
                "context_precision":r.ragas.get("context_precision",""),
                "context_recall":r.ragas.get("context_recall",""),
                "rouge1_f":r.rouge.get("rouge1",{}).get("f",""),
                "rouge2_f":r.rouge.get("rouge2",{}).get("f",""),
                "rougeL_f":r.rouge.get("rougeL",{}).get("f",""),
                "bertscore_f":r.bertscore.get("f1",""),
                "hallucination_rate":r.hallucination.get("hallucination_rate",""),
                "grounding_score":r.hallucination.get("grounding_score",""),
                "citation_accuracy":r.citations.get("accuracy",""),
                "n_citations":r.citations.get("n_citations",0),
                "latency_ms":r.latency_ms,
                "insufficient":r.insufficient,
                "failure_modes":"|".join(r.failure_modes),
                "question":r.question[:80],
                "answer_preview":r.answer[:120],
            })
    print(f"  [SAVED] {rc}")

    # Benchmark CSV
    rb = out_dir/f"ragas_benchmark_table_{model}.csv"
    rows = [["Metric","Score","Threshold"],
        ["-- RAGAS --","",""],
        ["Faithfulness",      f"{summary.faithfulness:.4f}",      ">=0.80 excellent"],
        ["Answer Relevancy",  f"{summary.answer_relevancy:.4f}",  ">=0.80 excellent"],
        ["Context Precision", f"{summary.context_precision:.4f}", ">=0.70 good"],
        ["Context Recall",    f"{summary.context_recall:.4f}",    ">=0.70 good"],
        ["-- ROUGE --","",""],
        ["ROUGE-1 F1", f"{summary.rouge1_f:.4f}", ">=0.30 reasonable"],
        ["ROUGE-2 F1", f"{summary.rouge2_f:.4f}", ">=0.15 reasonable"],
        ["ROUGE-L F1", f"{summary.rougeL_f:.4f}", ">=0.25 reasonable"],
        ["-- BERTScore --","",""],
        ["BERTScore F1",f"{summary.bertscore_f:.4f}",">=0.85 high similarity"],
        ["BERTScore P", f"{summary.bertscore_p:.4f}",""],
        ["BERTScore R", f"{summary.bertscore_r:.4f}",""],
        ["-- Hallucination --","",""],
        ["Hallucination Rate",f"{summary.avg_hallucination_rate:.4f}","<=0.20 acceptable"],
        ["Grounding Score",   f"{summary.avg_grounding_score:.4f}",   ">=0.80 good"],
        ["Citation Accuracy", f"{summary.avg_citation_accuracy or 0:.4f}",">=0.80 good"],
        ["-- System --","",""],
        ["N Questions",str(summary.n_evaluated),""],
        ["Insufficient Rate",f"{summary.n_insufficient/summary.n_total:.2%}",""],
        ["Est. Cost USD",f"${summary.estimated_cost_usd:.3f}",""],
        [],
        ["-- Per Category --"],
        ["Category","N","Faithfulness","Relevancy","ROUGE-1","BERTScore","Hall%"],
    ]
    for cat, d in summary.by_category.items():
        rows.append([cat, d["n"],
            f"{d.get('faithfulness',0):.4f}", f"{d.get('answer_relevancy',0):.4f}",
            f"{d.get('rouge1_f',0):.4f}",     f"{d.get('bertscore_f',0):.4f}",
            f"{d.get('hallucination_rate',0)*100:.1f}%"])
    with open(rb,"w",newline="",encoding="utf-8-sig") as f:
        csv.writer(f).writerows(rows)
    print(f"  [SAVED] {rb}")

    # Failure analysis JSON
    failures = defaultdict(list)
    for r in results:
        for fm in r.failure_modes:
            if fm != "NO_FAILURE":
                failures[fm].append({"qid":r.qid,"category":r.category,
                    "question":r.question[:80],"answer":r.answer[:120],
                    "ragas":r.ragas,"rouge1_f":r.rouge.get("rouge1",{}).get("f",0)})
    rf = out_dir/f"ragas_failure_analysis_{model}.json"
    with open(rf,"w",encoding="utf-8") as f:
        json.dump(dict(failures),f,indent=2,ensure_ascii=False,default=str)
    print(f"  [SAVED] {rf}")

    # Excel
    _save_excel(summary, results, out_dir, model)

    # Thesis text
    thesis = _thesis_text(summary, model)
    rt = out_dir/f"ragas_thesis_section_{model}.txt"
    with open(rt,"w",encoding="utf-8") as f: f.write(thesis)
    print(f"  [SAVED] {rt}")
    print(thesis)


def _save_excel(summary, results, out_dir, model):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("  [WARN] openpyxl not installed - skipping Excel"); return
    wb = openpyxl.Workbook()
    DARK="1E293B"; TEAL="0F766E"; WHITE="FFFFFF"; LGREY="F8FAFC"
    LGRN="DCFCE7"; LRED="FEE2E2"; LBLU="DBEAFE"; LAMB="FEF3C7"; LTEAL="CCFBF1"
    BORD="CBD5E1"
    thin=lambda: Side(style="thin",color=BORD)
    bdr=lambda: Border(left=thin(),right=thin(),top=thin(),bottom=thin())
    def hdr(ws,r,c,v,bg=DARK,size=9):
        cell=ws.cell(r,c,v); cell.font=Font(name="Calibri",bold=True,size=size,color=WHITE)
        cell.fill=PatternFill("solid",fgColor=bg)
        cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
        cell.border=bdr(); return cell
    def dat(ws,r,c,v,bg=LGREY,bold=False,align="center",size=9):
        cell=ws.cell(r,c,v); cell.font=Font(name="Calibri",bold=bold,size=size)
        cell.fill=PatternFill("solid",fgColor=bg)
        cell.alignment=Alignment(horizontal=align,vertical="center")
        cell.border=bdr(); return cell

    # Sheet 1: Summary
    ws=wb.active; ws.title="Summary"
    ws.merge_cells(f"A1:D1")
    c=ws.cell(1,1,f"End-to-End RAG Evaluation -- {model.upper()} -- n={summary.n_evaluated}")
    c.font=Font(name="Calibri",bold=True,size=12,color=WHITE)
    c.fill=PatternFill("solid",fgColor=DARK)
    c.alignment=Alignment(horizontal="center",vertical="center")
    ws.row_dimensions[1].height=24
    sections=[
        ("RAGAS (LLM-as-Judge)",[
            ("Faithfulness",summary.faithfulness,LTEAL,0.80,True),
            ("Answer Relevancy",summary.answer_relevancy,LTEAL,0.80,True),
            ("Context Precision",summary.context_precision,LTEAL,0.70,True),
            ("Context Recall",summary.context_recall,LTEAL,0.70,True)]),
        ("ROUGE vs Gold",[
            ("ROUGE-1 F1",summary.rouge1_f,LBLU,0.30,True),
            ("ROUGE-2 F1",summary.rouge2_f,LBLU,0.15,True),
            ("ROUGE-L F1",summary.rougeL_f,LBLU,0.25,True)]),
        ("BERTScore vs Gold",[
            ("BERTScore F1",summary.bertscore_f,LGRN,0.85,True),
            ("BERTScore P",summary.bertscore_p,LGRN,0.85,True),
            ("BERTScore R",summary.bertscore_r,LGRN,0.85,True)]),
        ("Hallucination",[
            ("Hall. Rate",summary.avg_hallucination_rate,LRED,0.20,False),
            ("Grounding Score",summary.avg_grounding_score,LGRN,0.80,True),
            ("Citation Accuracy",summary.avg_citation_accuracy or 0,LAMB,0.80,True)]),
    ]
    row=2; hdr(ws,row,1,"Metric",bg=TEAL); hdr(ws,row,2,"Score",bg=TEAL)
    hdr(ws,row,3,"Pass?",bg=TEAL); hdr(ws,row,4,"Threshold",bg=TEAL); row+=1
    for sec,rows in sections:
        ws.merge_cells(f"A{row}:D{row}")
        c=ws.cell(row,1,sec); c.font=Font(name="Calibri",bold=True,size=9,color=WHITE)
        c.fill=PatternFill("solid",fgColor=TEAL)
        c.alignment=Alignment(horizontal="left",vertical="center",indent=1)
        ws.row_dimensions[row].height=16; row+=1
        for metric,val,bg_ok,thresh,higher_better in rows:
            good=(val>=thresh if higher_better else val<=thresh)
            bg=LGRN if good else LRED
            dat(ws,row,1,metric,align="left",bg=LGREY,bold=True)
            dat(ws,row,2,round(val or 0,4),bg=bg,bold=True)
            dat(ws,row,3,"PASS" if good else "REVIEW",bg=bg)
            dat(ws,row,4,f"{'>' if higher_better else '<'}={thresh}",align="left",bg=LGREY)
            row+=1
    ws.column_dimensions["A"].width=24; ws.column_dimensions["B"].width=10
    ws.column_dimensions["C"].width=10; ws.column_dimensions["D"].width=18

    # Sheet 2: Per-query
    ws2=wb.create_sheet("Per-Query")
    cols=["ID","Cat","Diff","Faith","Rel","C.Pr","C.Rc","R1","R2","RL","BSCO","Hall%","Grnd","Insuf","Failures","Question"]
    for j,col in enumerate(cols,1): hdr(ws2,1,j,col)
    ws2.row_dimensions[1].height=22
    CCAT={"factual":LBLU,"comparative":LAMB,"gap_detection":LTEAL}
    for i,r in enumerate(results,2):
        bg=CCAT.get(r.category,LGREY)
        faith=r.ragas.get("faithfulness",0.5)
        qbg=LGRN if faith>=0.8 else (LAMB if faith>=0.5 else LRED)
        for j,v in enumerate([
            r.qid,r.category[:4],r.difficulty[:3],
            round(r.ragas.get("faithfulness",0),3),
            round(r.ragas.get("answer_relevancy",0),3),
            round(r.ragas.get("context_precision",0),3),
            round(r.ragas.get("context_recall",0),3),
            round(r.rouge.get("rouge1",{}).get("f",0),3),
            round(r.rouge.get("rouge2",{}).get("f",0),3),
            round(r.rouge.get("rougeL",{}).get("f",0),3),
            round(r.bertscore.get("f1",0),3),
            f"{r.hallucination.get('hallucination_rate',0)*100:.0f}%",
            round(r.hallucination.get("grounding_score",0),3),
            "YES" if r.insufficient else "no",
            "|".join(r.failure_modes),
            r.question[:60],
        ],1):
            dat(ws2,i,j,v,bg=qbg if j<=14 else LGREY,
                align="left" if j>=15 else "center",size=8)
    for j,w in enumerate([8,5,5,6,6,6,6,6,6,6,6,6,6,5,20,42],1):
        ws2.column_dimensions[get_column_letter(j)].width=w

    # Sheet 3: Category
    ws3=wb.create_sheet("Category Breakdown")
    cat_cols=["Category","N","Faith.","Rel.","Ctx.Pr","Ctx.Rc","ROUGE-1","BSCO","Hall%","Lat(ms)"]
    for j,col in enumerate(cat_cols,1): hdr(ws3,1,j,col)
    r3=2
    for cat,d in summary.by_category.items():
        cbg=CCAT.get(cat,LGREY)
        for j,v in enumerate([cat,d["n"],
            f"{d.get('faithfulness',0):.4f}",f"{d.get('answer_relevancy',0):.4f}",
            f"{d.get('context_precision',0):.4f}",f"{d.get('context_recall',0):.4f}",
            f"{d.get('rouge1_f',0):.4f}",f"{d.get('bertscore_f',0):.4f}",
            f"{d.get('hallucination_rate',0)*100:.1f}%",
            f"{d.get('avg_latency_ms',0):.0f}"],1):
            dat(ws3,r3,j,v,bg=cbg,align="left" if j==1 else "center",size=9)
        r3+=1
    for j,w in enumerate([18,5,10,10,10,10,10,10,10,10],1):
        ws3.column_dimensions[get_column_letter(j)].width=w

    path=out_dir/f"ragas_benchmark_table_{model}.xlsx"
    wb.save(path); print(f"  [SAVED] {path}")


def _thesis_text(s: EvalSummary, model: str) -> str:
    faith_q=("excellent" if s.faithfulness>=0.80 else "good" if s.faithfulness>=0.60 else "moderate")
    hall_q =("low (acceptable)" if s.avg_hallucination_rate<=0.20 else
             "moderate" if s.avg_hallucination_rate<=0.35 else "high (concerning)")
    return f"""{'='*65}
  THESIS CH.5 SECTION 5.3 -- RAG EVALUATION RESULTS
{'='*65}

Section 5.3  End-to-End RAG Evaluation

We evaluated the full RAG pipeline on {s.n_evaluated} questions from
the golden evaluation dataset using three complementary frameworks:
RAGAS (Es et al., 2024), ROUGE (Lin, 2004), and BERTScore
(Zhang et al., 2020). The {model.upper()} backend with top-k={s.top_k}
retrieved chunks was used for all experiments.

Table X -- End-to-End RAG Evaluation Results ({model.upper()})

  RAGAS Metrics (LLM-as-Judge, GPT-4o)
    Faithfulness      = {s.faithfulness:.4f}   [{faith_q}]
    Answer Relevancy  = {s.answer_relevancy:.4f}
    Context Precision = {s.context_precision:.4f}
    Context Recall    = {s.context_recall:.4f}

  Lexical Overlap vs Gold (ROUGE)
    ROUGE-1 F1        = {s.rouge1_f:.4f}
    ROUGE-2 F1        = {s.rouge2_f:.4f}
    ROUGE-L F1        = {s.rougeL_f:.4f}

  Semantic Similarity vs Gold (BERTScore)
    BERTScore F1      = {s.bertscore_f:.4f}
    BERTScore P       = {s.bertscore_p:.4f}
    BERTScore R       = {s.bertscore_r:.4f}

  Hallucination Analysis (per-claim GPT-4o verification)
    Hallucination Rate  = {s.avg_hallucination_rate:.4f}  [{hall_q}]
    Grounding Score     = {s.avg_grounding_score:.4f}
    Citation Accuracy   = {s.avg_citation_accuracy or 0:.4f}

Faithfulness of {s.faithfulness:.4f} indicates that the majority of
claims in generated answers are {"grounded in" if s.faithfulness>=0.6 else "not fully grounded in"}
retrieved context. The per-claim hallucination rate of
{s.avg_hallucination_rate:.4f} ({hall_q}) was estimated via GPT-4o
verification of each extracted claim against its source chunks.
This is consistent with Dahl et al. (2024, Stanford) who find
hallucination rates of 17-58% in commercial legal RAG systems,
situating our system's performance in the broader policy-domain
RAG literature.

ROUGE-1 F1 of {s.rouge1_f:.4f} reflects lexical overlap between
system answers and gold references. ROUGE scores for open-ended
policy QA are inherently lower than for extractive tasks, as
generative systems express correct information using different
wording (Lin, 2004). BERTScore F1 of {s.bertscore_f:.4f} captures
semantic equivalence that ROUGE misses -- answers paraphrasing
"carbon neutrality" as "net zero" will score well on BERTScore
but poorly on ROUGE.

Table Y -- Per-Category Breakdown
  Category           Faith.    Rel.    ROUGE-1  BERTScore  Hall%
{chr(10).join(
  f"  {cat:<18} {d.get('faithfulness',0):.4f}  {d.get('answer_relevancy',0):.4f}  {d.get('rouge1_f',0):.4f}  {d.get('bertscore_f',0):.4f}  {d.get('hallucination_rate',0)*100:.1f}%"
  for cat, d in s.by_category.items()
)}

Est. API cost for this evaluation: ~${s.estimated_cost_usd:.3f} USD
"""


# =============================================================================
#  SECTION 11 -- CLI ENTRY POINT
# =============================================================================

def main():
    p = argparse.ArgumentParser(
        description="End-to-end RAG evaluation: RAGAS + ROUGE + BERTScore")
    p.add_argument("--base_dir",         default=".")
    p.add_argument("--model",            default="gpt4o", choices=["gpt4o","mistral_api"])
    p.add_argument("--top_k",            default=5, type=int)
    p.add_argument("--n_questions",      default=80, type=int)
    p.add_argument("--category",         default="all",
                   choices=["all","factual","comparative","gap_detection"])
    p.add_argument("--difficulty",       default="all",
                   choices=["all","easy","medium","hard"])
    p.add_argument("--golden_dataset",   default="Evaluation_Code/eval_golden_dataset.json")
    p.add_argument("--out_dir",          default="data/eval")
    p.add_argument("--skip_ragas",       action="store_true")
    p.add_argument("--skip_bertscore",   action="store_true")
    p.add_argument("--skip_hallucination", action="store_true")
    p.add_argument("--verbose",          action="store_true", default=True)
    args = p.parse_args()

    base_dir    = Path(args.base_dir).resolve()
    golden_path = (base_dir/args.golden_dataset).resolve()
    out_dir     = (base_dir/args.out_dir).resolve()

    # Find golden dataset
    if not golden_path.exists():
        for alt in [base_dir/"eval_golden_dataset.json",
                    Path(__file__).parent/"eval_golden_dataset.json"]:
            if alt.exists(): golden_path=alt; break
    if not golden_path.exists():
        print(f"ERROR: Golden dataset not found at {golden_path}"); sys.exit(1)

    api_key = os.environ.get("OPENAI_API_KEY")
    if not api_key:
        print("ERROR: OPENAI_API_KEY not set."); sys.exit(1)

    with open(golden_path, encoding="utf-8") as f:
        all_qs = json.load(f)

    questions = [q for q in all_qs
                 if (args.category=="all" or q["category"]==args.category)
                 and (args.difficulty=="all" or q["difficulty"]==args.difficulty)
                 ][:args.n_questions]

    print(f"\n{chr(45)*60}")
    print(f"  End-to-End RAG Evaluation")
    print(f"  Model:     {args.model}")
    print(f"  Questions: {len(questions)}/{len(all_qs)}")
    print(f"  RAGAS:     {'SKIPPED' if args.skip_ragas else 'YES'}")
    print(f"  BERTScore: {'SKIPPED' if args.skip_bertscore else 'YES'}")
    print(f"  Hall.:     {'SKIPPED' if args.skip_hallucination else 'YES'}")
    if not args.skip_ragas and len(questions) >= 20:
        est = len(questions)*0.015
        print(f"  [COST] ~{len(questions)*3} GPT-4o calls estimated ~${est:.2f}")
    print(f"{chr(45)*60}\n")

    evaluator = RAGEvaluator(
        base_dir=base_dir, api_key=api_key, model=args.model,
        top_k=args.top_k, skip_ragas=args.skip_ragas,
        skip_bertscore=args.skip_bertscore,
        skip_hallucination=args.skip_hallucination,
        verbose=args.verbose,
    )

    from generation import RAGPipeline
    pipe = RAGPipeline(backend=args.model, top_k=args.top_k,
                       embed_device="cpu", openai_key=api_key)

    t0 = time.time()
    results = evaluator.collect(questions, pipe)
    results = evaluator.rouge_pass(results)
    if not args.skip_ragas:
        print("\n  [Step 3] Running RAGAS...")
        results = run_ragas(results, api_key)
    else:
        print("\n  [Step 3] RAGAS SKIPPED.")
    if not args.skip_bertscore:
        print("\n  [Step 4] Running BERTScore...")
        results = run_bertscore(results)
    else:
        print("\n  [Step 4] BERTScore SKIPPED.")
    results = evaluator.hallucination_pass(results)
    results = evaluator.failure_pass(results)
    summary = evaluator.aggregate(results)
    print(f"\n  Total time: {time.time()-t0:.0f}s")
    print_summary(summary)
    print("\n  Saving outputs...")
    save_results(summary, results, out_dir, args.model)
    print(f"\n  Done. Outputs in {out_dir}")

if __name__ == "__main__":
    main()
